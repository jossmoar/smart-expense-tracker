import io
from abc import ABC, abstractmethod
from typing import List

from app.models.expense import Expense


class ReportGenerator(ABC):
    @abstractmethod
    def generate(self, expenses: List[Expense]) -> bytes: ...


class PDFReportGenerator(ReportGenerator):
    def generate(self, expenses: List[Expense]) -> bytes:
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas

        buffer = io.BytesIO()
        pdf = canvas.Canvas(buffer, pagesize=letter)
        pdf.setFont("Helvetica-Bold", 14)
        pdf.drawString(40, 760, "Reporte de gastos")

        pdf.setFont("Helvetica", 10)
        y = 730
        total = 0.0
        for expense in expenses:
            pdf.drawString(40, y, f"{expense.date}  {expense.category:<15}  ₡{expense.amount:,.0f}")
            total += expense.amount
            y -= 16
            if y < 40:
                pdf.showPage()
                y = 760

        pdf.setFont("Helvetica-Bold", 11)
        pdf.drawString(40, y - 10, f"Total: ₡{total:,.0f}")
        pdf.save()
        return buffer.getvalue()


class ExcelReportGenerator(ReportGenerator):
    def generate(self, expenses: List[Expense]) -> bytes:
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Gastos"
        sheet.append(["Fecha", "Categoría", "Descripción", "Monto"])

        for expense in expenses:
            sheet.append([str(expense.date), expense.category, expense.description, expense.amount])
            sheet.cell(row=sheet.max_row, column=4).number_format = '"₡"#,##0'

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()


class ReportFactory:
    """Factory Method: the caller asks for a format, not a concrete class —
    adding a new export type (CSV, JSON) means one new branch here, nothing
    else changes."""

    _generators = {
        "pdf": PDFReportGenerator,
        "excel": ExcelReportGenerator,
    }

    @classmethod
    def create(cls, report_format: str) -> ReportGenerator:
        generator_cls = cls._generators.get(report_format)
        if generator_cls is None:
            raise ValueError(f"Unsupported report format: {report_format}")
        return generator_cls()
